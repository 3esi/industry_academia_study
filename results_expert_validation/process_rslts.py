import pandas as pd
import os
from openpyxl import Workbook
import re
from typing import Dict, Any
from pathlib import Path

def transpose_excel_file(df):
    """
    Transpose function that converts a 2-column Excel into Question-Answer format,
    or transposes any Excel and renames first two columns to Question-Answer.
    
    Parameters:
    -----------
    input_file : str
        Path to the input Excel file
    output_file : str, optional
        Path to the output Excel file. If None, creates a new file with '_qa' suffix
    
    Returns:
    --------
    pd.DataFrame
        DataFrame with 'Questions' and 'Answers' columns
    """
    output_file = "csv_files/results_expers_transposed.xlsx"
    sheet_name=0

    try:
        # Read the Excel file WITHOUT setting any column as index and WITHOUT header
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
        
        # Transpose the DataFrame
        df_transposed = df.transpose()
        
        # Take only the first two columns and rename them
        if df_transposed.shape[1] >= 2:
            qa_df = df_transposed.iloc[:, :2].copy()
            qa_df.columns = ['Questions', 'Answers']
        elif df_transposed.shape[1] == 1:
            qa_df = df_transposed.copy()
            qa_df.columns = ['Questions']
            qa_df['Answers'] = ''
        else:
            print("Error: No data to transpose")
            return None
        
        # Reset index to make it cleaner
        qa_df = qa_df.reset_index(drop=True)
        
        # Generate output filename if not provided
        if output_file is None:
            file_name, file_ext = os.path.splitext(input_file)
            output_file = f"{file_name}_qa{file_ext}"
        
        # Save the Q&A DataFrame to Excel
        qa_df.to_excel(output_file, index=False)
        
        print(f"Q&A file saved as: {output_file}")
        
        return qa_df
        
    except FileNotFoundError:
        print(f"Error: File '{input_file}' not found.")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def extract_paper_title(cell_value):
    """
    Extracts the paper title from a cell that contains the title of a paper.
    Returns only the title of the paper.
    
    Parameters:
    cell_value (str): The full cell value
    
    Returns:
    str: The cleaned paper title
    """
    
    cell_value = str(cell_value).strip()
    
    paper_title_start = cell_value.lower().find('paper title:')
    if paper_title_start == -1:
        return cell_value  # If no "Paper title:" found, return original
    
    # Extract everything after "Paper title:"
    after_paper_title = cell_value[paper_title_start + len('paper title:'):].strip()
    
    # Find "Introductory text" (case insensitive)
    intro_text_start = after_paper_title.lower().find('introductory text')
    if intro_text_start != -1:
        title_only = after_paper_title[:intro_text_start].strip()
    else:
        print(f"No intro text in helper function to clean the paper title.")
    
    return title_only

def extract_title_and_answers(df, paper_title_positions):
    """
    Extracts the answers for all questions related to a paper title (i.e. the next 6 lines following a paper title).
    
    Parameters:
    df (pandas.DataFrame): The rows and columsn of the original excel sheet
    paper_title_positions (list): Positions of type int of the paper titles in df
    
    Returns:
    list: The paper title and the answers to the questions
    """
    extracted_records = []

    # Extract data for each "Paper title" found
    for row_idx in paper_title_positions:
        # Check if we have enough rows after the "Paper title" row
        if row_idx + 6 < len(df):
            record_data = []
            
            # First, get the paper title from Questions column
            paper_title_value = df['Questions'].iloc[row_idx]
            if pd.isna(paper_title_value):
                paper_title_value = ''
            else:
                paper_title_value = extract_paper_title(paper_title_value)
            record_data.append(paper_title_value)
            
            # Then get the next 6 values from the Answers column
            for i in range(1, 7):
                current_row = row_idx + i
                cell_value = df['Answers'].iloc[current_row]
                
                # Handle NaN values
                if pd.isna(cell_value):
                    cell_value = ''
                else:
                    cell_value = str(cell_value).strip()
                
                record_data.append(cell_value)
            
            # Add the record to our list
            extracted_records.append(record_data)
        
        else:
            print(f"Warning: Not enough rows after 'Paper title' at row {row_idx}")

    return extracted_records
    
def clean_excel_data(input_file, output_file, sheet_name=0):
    """
    Extracts data from Excel file where 'Paper title' appears and saves
    the following 6 rows from the 'Answers' column as structured data.
    
    Parameters:
    input_file (str): Path to input Excel file
    output_file (str): Path to output Excel file
    sheet_name (str or int): Sheet name or index to read from

    Returns:
    pandas.DataFrame: The cleaned rows and columns of the original excel sheet
    """
    
    # Define output columns
    column_names = [
        'Paper_Title',
        'Understanding (self)',
        'Understanding (others)',
        'Level of detail',
        'Additional info',
        'Distracting info',
        'Comments'
    ]
    
    try:
        df_notclean = pd.read_excel(input_file, sheet_name=sheet_name)

        # transpose the excel file for better clarity
        df = transpose_excel_file(df_notclean)

        # Check if required columns exist
        required_columns = ['Questions', 'Answers']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"Error: Missing columns: {missing_columns}")
            print(f"Available columns: {list(df.columns)}")
            return None
        
        extracted_records = []
        
        # Convert Questions column to string for searching, handle NaN values
        questions_series = df['Questions'].astype(str).fillna('')
        
        # Search for rows containing "Paper title" in the Questions column
        paper_title_positions = []
        for row_idx in range(len(questions_series)):
            cell_value = questions_series.iloc[row_idx].lower()
            if 'paper title' in cell_value:
                paper_title_positions.append(row_idx)

        # Extract paper title and answers to the questions for this paper title
        extracted_records = extract_title_and_answers(df, paper_title_positions)
        
        # Create DataFrame with extracted records
        if extracted_records:
            output_df = pd.DataFrame(extracted_records, columns=column_names)
            
            # Clean up the data (remove extra whitespace, handle empty cells)
            for col in output_df.columns:
                output_df[col] = output_df[col].astype(str).str.strip()
                output_df[col] = output_df[col].replace('nan', '')
            
            # Save to Excel file
            output_df.to_excel(output_file, index=False)
            print(f"Successfully saved {len(output_df)} records to {output_file}")
            
            return output_df
            
        else:
            print("No records found containing 'Paper title'")
            # Create empty DataFrame with correct columns
            empty_df = pd.DataFrame(columns=column_names)
            empty_df.to_excel(output_file, index=False)
            return empty_df
            
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found")
        return None
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        return None

def create_combined_excel(input_folder):
    """
    Creates an Excel file called 'input_results_experts_all.xlsx' by combining
    Excel files from the 'input' folder. Each file becomes a separate sheet
    named P1, P2, P3, etc., based on the file's suffix (p1, p2, p3, etc.).
    """
    output_filename = "csv_files/input_results_experts_all.xlsx"
    
    # Check if input folder exists
    if not os.path.exists(input_folder):
        print(f"Error: '{input_folder}' folder not found!")
        return
    
    # Get all Excel files from the input folder
    excel_files = []
    for file in os.listdir(input_folder):
        if file.endswith(('.xlsx', '.xls')) and 'results_experts_processed_p' in file:
            excel_files.append(file)
    
    if not excel_files:
        print("No matching Excel files found in the input folder!")
        return
    
    # Sort files to ensure consistent ordering
    excel_files.sort()
    
    # Create a dictionary to store dataframes with their sheet names
    sheets_data = {}
    
    for file in excel_files:
        file_path = os.path.join(input_folder, file)
        
        # Extract the number from filename using regex
        # Matches pattern like "results_experts_processed_p3.xlsx"
        match = re.search(r'results_experts_processed_p(\d+)', file)
        
        if match:
            number = match.group(1)
            sheet_name = f"P{number}"
            
            try:
                # Read the Excel file
                df = pd.read_excel(file_path)
                sheets_data[sheet_name] = df
                print(f"Loaded {file} as sheet '{sheet_name}'")
                
            except Exception as e:
                print(f"Error reading {file}: {str(e)}")
                continue
        else:
            print(f"Warning: Could not extract number from filename '{file}'")
    
    if not sheets_data:
        print("No valid data found to combine!")
        return
    
    # Write all sheets to a single Excel file
    try:
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"\nSuccessfully created '{output_filename}' with {len(sheets_data)} sheets:")
        for sheet_name in sorted(sheets_data.keys()):
            print(f"  - {sheet_name}")
            
    except Exception as e:
        print(f"Error writing output file: {str(e)}")

def consolidate_sheets(input_file_path, sheet_names=None):
    """
    Compare three Excel sheets and create a NEW Excel file with Results.
    This avoids any issues with modifying existing files.
    
    Args:
        input_file_path (str): Path to the input Excel file
        output_file_path (str): Path for the new output file. If None, adds '_results' to input filename
        sheet_names (list): List of 3 sheet names to compare. If None, uses first 3 sheets.
    
    Returns:
        str: Status message
    """
    
    try:
        output_file_path = "csv_files/results_experts_consolidated.xlsx"
        
        # Read the sheets using pandas (skip first row only)
        if sheet_names is None:
            excel_file = pd.ExcelFile(input_file_path)
            sheet_names = excel_file.sheet_names[:3]
        
        if len(sheet_names) < 3:
            raise ValueError("Error: Need at least 3 sheets to compare")
        
        df1 = pd.read_excel(input_file_path, sheet_name=sheet_names[0], header=None).iloc[1:, :]
        df2 = pd.read_excel(input_file_path, sheet_name=sheet_names[1], header=None).iloc[1:, :]
        df3 = pd.read_excel(input_file_path, sheet_name=sheet_names[2], header=None).iloc[1:, :]
        
        # Find maximum dimensions
        max_rows = max(len(df1), len(df2), len(df3))
        max_cols = max(df1.shape[1], df2.shape[1], df3.shape[1])
        
        # Create new workbook
        wb = Workbook()
        
        # Copy original sheets into new workbook (keeping first row skipped)
        for sheet_name in sheet_names:
            ws = wb.create_sheet(title=sheet_name)
            df = pd.read_excel(input_file_path, sheet_name=sheet_name, header=None).iloc[1:, :]
            
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row):
                    if pd.notna(value):
                        ws.cell(row=row_idx+1, column=col_idx+1, value=value)
        
        # Remove default sheet
        wb.remove(wb['Sheet'])
        
        # Create Results sheet
        results_ws = wb.create_sheet("Results", 0)  # Insert as first sheet
        
        # Define new headers
        headers = [
            "Paper_Title",
            "Understanding (self)",
            "Understanding (others)",
            "Level of detail",
            "Additional info",
            "Distracting info",
            "Comments"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            results_ws.cell(row=1, column=col_idx, value=header)
        
        # Compare and fill Results sheet
        for row in range(max_rows):
            for col in range(min(max_cols, len(headers))):
                # Get values from each dataframe
                val1 = df1.iloc[row, col] if row < len(df1) and col < df1.shape[1] else None
                val2 = df2.iloc[row, col] if row < len(df2) and col < df2.shape[1] else None
                val3 = df3.iloc[row, col] if row < len(df3) and col < df3.shape[1] else None
                
                # Handle NaN values
                if pd.isna(val1): val1 = None
                if pd.isna(val2): val2 = None
                if pd.isna(val3): val3 = None
                
                formatted_val1 = "" if val1 is None else str(val1)
                formatted_val2 = "" if val2 is None else str(val2)
                formatted_val3 = "" if val3 is None else str(val3)
                
                if col == 0:  
                    # Special rule for Paper_Title
                    if formatted_val1 == formatted_val2 == formatted_val3:
                        combined_value = formatted_val1  # all same → just keep one
                    else:
                        combined_value = f"P1: {formatted_val1}, P2: {formatted_val2}, P3: {formatted_val3}"
                else:
                    # Default behavior for other columns
                    combined_value = f"P1: {formatted_val1}, P2: {formatted_val2}, P3: {formatted_val3}"
                
                results_ws.cell(row=row+2, column=col+1, value=combined_value)  # +2 for headers
        
        # Save the new workbook
        wb.save(output_file_path)
        wb.close()
        
        return f"Success! Created new file '{output_file_path}'. Compared sheets: {', '.join(sheet_names)}"
    
    except Exception as e:
        return f"Error: {str(e)}"
    
if __name__ == "__main__":
    input_file = "results_survey/input_results_experts_p3.xlsx"
    output_file = "results_survey/results_experts_processed_p3.xlsx"

    # file(s) after basic data cleaning
    input_files_cleaned_all = "csv_files/input_results_experts_all.xlsx"

    # file(s) after consolidation
    input_file_consolidated = "csv_files/results_experts_consolidated.xlsx" 

    if os.path.exists(input_files_cleaned_all):
        print(f"{input_files_cleaned_all} exists! Consolidating cleaned results.")
        result = consolidate_sheets(input_files_cleaned_all)

        print(result)
    else:
        # Perform basic data cleaning if not performed yet
        print(f"{input_files_cleaned_all} does not exist! Initiating data cleaning.")
        result_df = clean_excel_data(input_file, output_file)

        input_folder = "results_survey"
        create_combined_excel(input_folder)
    
        if result_df is not None:
            print("Data cleaning completed successfully!")
        else:
            print("Data cleaning failed.")

    

    

    